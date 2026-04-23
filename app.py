#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║         Job Seeker Prompt Generator - Web Version      ║
║  Excel + Resume → Ready-to-paste AI prompts             ║
╚══════════════════════════════════════════════════════════╝

Web interface for the Job Seeker Prompt Generator.
Upload your Excel file and resume files, then generate AI prompts
for job applications.

No API key required. Generate prompts for ChatGPT, Claude, Gemini,
or any other AI interface.
"""

from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session
import os
import tempfile
import zipfile
import uuid
import shutil
from io import BytesIO
from datetime import datetime
from pathlib import Path
import openpyxl
import json

app = Flask(__name__)
app.secret_key = 'job-search-generator-secret-key'

# Create temporary directory for downloads
TEMP_DIR = Path(tempfile.gettempdir()) / 'job_search_downloads'
TEMP_DIR.mkdir(exist_ok=True)

def cleanup_old_sessions():
    """Clean up temporary files older than 1 hour."""
    import time
    current_time = time.time()
    max_age = 3600  # 1 hour

    for session_dir in TEMP_DIR.iterdir():
        if session_dir.is_dir():
            try:
                # Check if directory is older than max_age
                dir_age = current_time - session_dir.stat().st_mtime
                if dir_age > max_age:
                    shutil.rmtree(session_dir)
                    print(f"Cleaned up old session: {session_dir.name}")
            except Exception as e:
                print(f"Error cleaning up {session_dir}: {e}")

# Clean up old sessions on startup
cleanup_old_sessions()

# ─────────────────────────── RUBRIC ────────────────────────────────────────

RUBRIC = [
    {
        "id": 1,
        "rule": "Specificity",
        "description": (
            "The response must reference specific details from the job description, "
            "company name, or resume - not generic advice."
        ),
    },
    {
        "id": 2,
        "rule": "Actionability",
        "description": (
            "Every answer must include at least one concrete, actionable takeaway "
            "the candidate can act on immediately."
        ),
    },
    {
        "id": 3,
        "rule": "Completeness",
        "description": (
            "The response must fully address the question asked without leaving "
            "major sub-points unanswered."
        ),
    },
    {
        "id": 4,
        "rule": "Candidate-Centric Framing",
        "description": (
            "Insights must be framed from the job seeker's perspective - helping "
            "them prepare, position themselves, or improve."
        ),
    },
    {
        "id": 5,
        "rule": "Professional Tone",
        "description": (
            "Language must be clear, professional, and free of filler phrases like "
            "'Great question!' or vague platitudes."
        ),
    },
]

# ─────────────────────────── PROMPTS ───────────────────────────────────────

def build_prompts(
    company: str,
    job_desc: str,
    resume_summary: str,
    resume: str,
) -> list:
    """Build the five prompts with real data injected."""
    return [
        {
            "id":    "culture",
            "label": "1. Company Culture & Intel",
            "instructions": (
                "Paste this into your AI of choice. It will summarize the company's "
                "culture, known challenges, and recent news to help you prepare."
            ),
            "prompt": (
                f"Summarize {company}'s culture, challenges, and recent news "
                "for a job interview. Be specific about what a candidate should "
                "know walking into an interview. Include cultural values, known "
                "workplace challenges, recent news or leadership changes, and one "
                "smart question the candidate could ask that shows insider knowledge."
            ),
        },
        {
            "id":    "hidden_requirements",
            "label": "2. Hidden Job Requirements",
            "instructions": (
                "Paste this prompt to uncover the unstated priorities and personality "
                "traits the hiring manager actually wants."
            ),
            "prompt": (
                "What is this job actually looking for underneath the formal "
                "requirements? Read between the lines and surface the unstated "
                "priorities, pain points, and personality traits the hiring manager "
                "really wants.\n\n"
                f"--- JOB DESCRIPTION ---\n{job_desc}"
            ),
        },
        {
            "id":    "gap_analysis",
            "label": "3. Gap Analysis",
            "instructions": (
                "Paste this prompt to get an honest assessment of where your "
                "background is weakest for this role, with reframing suggestions."
            ),
            "prompt": (
                "Here is a job description and my background. Where am I weakest "
                "for this role? Be honest and specific. For each gap, suggest a "
                "concrete way to address or reframe it.\n\n"
                f"--- JOB DESCRIPTION ---\n{job_desc}\n\n"
                f"--- MY RESUME ---\n{resume}"
            ),
        },
        {
            "id":    "resume_summary",
            "label": "4. Tailored Resume Summary",
            "instructions": (
                "Paste this prompt to get a rewritten resume summary that mirrors "
                "the language and priorities of the job description."
            ),
            "prompt": (
                f"Rewrite my resume summary specifically for the role at {company}. "
                "Mirror the language and priorities in the job description. "
                "Keep it to 3-4 sentences, first-person, no fluff.\n\n"
                f"--- JOB DESCRIPTION ---\n{job_desc}\n\n"
                f"--- MY CURRENT SUMMARY ---\n{resume_summary}"
            ),
        },
        {
            "id":    "rubric_eval",
            "label": "5. Rubric Evaluation (paste AFTER getting a response)",
            "instructions": (
                "After the AI answers any of the above prompts, paste this as a "
                "follow-up to score the response against 5 quality rules."
            ),
            "prompt": (
                "Please evaluate your previous response against this rubric. "
                "For each rule give PASS or FAIL with one sentence of reasoning. "
                "End with an overall score out of 5 and one improvement suggestion.\n\n"
                "RUBRIC:\n"
                + "\n".join(
                    f"{r['id']}. {r['rule']}: {r['description']}"
                    for r in RUBRIC
                )
                + "\n\n"
                "Format:\n"
                "1. Specificity: [PASS/FAIL] - <reasoning>\n"
                "2. Actionability: [PASS/FAIL] - <reasoning>\n"
                "3. Completeness: [PASS/FAIL] - <reasoning>\n"
                "4. Candidate-Centric Framing: [PASS/FAIL] - <reasoning>\n"
                "5. Professional Tone: [PASS/FAIL] - <reasoning>\n\n"
                "Overall Score: X/5\n"
                "Improvement Suggestion: <one sentence>"
            ),
        },
    ]

# ─────────────────────────── EXCEL READER ──────────────────────────────────

def read_excel_from_upload(file) -> list:
    """Read Excel data from uploaded file object."""
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        required = {"Company", "Job_Description"}
        missing = required - set(headers)
        if missing:
            raise ValueError(
                f"Excel is missing required columns: {missing}\n"
                f"Found: {headers}"
            )

        col_map = {h: i for i, h in enumerate(headers)}
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company = row[col_map["Company"]]
            if not company:
                continue
            rows.append({
                "row":             row_idx,
                "company":         str(company).strip(),
                "job_description": str(row[col_map["Job_Description"]] or "").strip(),
            })

        if not rows:
            raise ValueError("No data rows found in the Excel file.")
        return rows
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")

# ─────────────────────────── OUTPUT BUILDERS ───────────────────────────────

def build_markdown(
    company: str,
    prompts: list,
    timestamp: str,
) -> str:
    """Build a clean markdown document with every prompt ready to copy."""

    lines = [
        f"# Job Application Prompts — {company}",
        f"",
        f"*Generated: {timestamp}*",
        f"",
        f"---",
        f"",
        f"## How to use this file",
        f"",
        f"Each section below contains one ready-to-paste prompt.",
        f"Copy the text inside the code block and paste it into",
        f"[ChatGPT](https://chat.openai.com), [Claude](https://claude.ai),",
        f"[Gemini](https://gemini.google.com), or any other AI web interface.",
        f"Run the prompts in order for best results.",
        f"",
        f"---",
        f"",
        f"## Quality Rubric",
        f"",
        f"Every response you receive should meet these 5 rules.",
        f"Use **Prompt 5** to ask the AI to self-evaluate.",
        f"",
        f"| # | Rule | What it checks |",
        f"|---|------|----------------|",
    ]

    for r in RUBRIC:
        lines.append(f"| {r['id']} | **{r['rule']}** | {r['description']} |")

    lines += ["", "---", ""]

    for item in prompts:
        lines += [
            f"## {item['label']}",
            f"",
            f"> {item['instructions']}",
            f"",
            f"```",
            item["prompt"],
            f"```",
            f"",
            f"---",
            f"",
        ]

    return "\n".join(lines)

# ─────────────────────────── ROUTES ────────────────────────────────────────

@app.route('/')
def index():
    """Main page with upload form."""
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    """Process uploaded files and generate prompts."""
    try:
        # Get uploaded files
        excel_file = request.files.get('excel_file')
        resume_file = request.files.get('resume_file')
        summary_file = request.files.get('summary_file')

        # Get text inputs (alternative to file uploads)
        resume_text = request.form.get('resume_text', '').strip()
        summary_text = request.form.get('summary_text', '').strip()

        # Validate inputs
        if not excel_file or excel_file.filename == '':
            flash('Please upload an Excel file with job data.', 'error')
            return redirect(url_for('index'))

        # Get resume content
        if resume_file and resume_file.filename != '':
            resume_content = resume_file.read().decode('utf-8').strip()
        elif resume_text:
            resume_content = resume_text
        else:
            flash('Please provide resume content either by uploading a file or pasting text.', 'error')
            return redirect(url_for('index'))

        # Get summary content
        if summary_file and summary_file.filename != '':
            summary_content = summary_file.read().decode('utf-8').strip()
        elif summary_text:
            summary_content = summary_text
        else:
            flash('Please provide resume summary content either by uploading a file or pasting text.', 'error')
            return redirect(url_for('index'))

        if not resume_content:
            flash('Resume content cannot be empty.', 'error')
            return redirect(url_for('index'))

        if not summary_content:
            flash('Resume summary content cannot be empty.', 'error')
            return redirect(url_for('index'))

        # Read Excel data
        rows = read_excel_from_upload(excel_file)

        # Generate prompts for all companies
        results = []
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

        for row in rows:
            company = row["company"]
            prompts = build_prompts(
                company=company,
                job_desc=row["job_description"],
                resume_summary=summary_content,
                resume=resume_content,
            )

            results.append({
                'company': company,
                'prompts': prompts,
                'markdown': build_markdown(company, prompts, timestamp)
            })

        # Save results to temporary files for download
        session_id = str(uuid.uuid4())
        session_dir = TEMP_DIR / session_id
        session_dir.mkdir(exist_ok=True)

        # Save individual markdown files
        for result in results:
            safe_company = "".join(
                c if c.isalnum() or c in "-_" else "_" for c in result['company']
            )
            filename = f"{safe_company}_prompts.md"
            filepath = session_dir / filename
            filepath.write_text(result['markdown'], encoding='utf-8')
            result['filename'] = filename

        # Save session info
        session_info = {
            'timestamp': timestamp,
            'companies': [r['company'] for r in results],
            'files': {r['company']: r['filename'] for r in results}
        }
        session_file = session_dir / 'session.json'
        session_file.write_text(json.dumps(session_info), encoding='utf-8')

        return render_template('results.html', results=results, timestamp=timestamp, session_id=session_id)

    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'An unexpected error occurred: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/download/<session_id>/<company>')
def download_single(session_id, company):
    """Download a single company's prompts as markdown."""
    try:
        session_dir = TEMP_DIR / session_id
        session_file = session_dir / 'session.json'

        if not session_file.exists():
            flash('Download session expired. Please generate prompts again.', 'error')
            return redirect(url_for('index'))

        session_info = json.loads(session_file.read_text(encoding='utf-8'))

        if company not in session_info['files']:
            flash('Company not found in this session.', 'error')
            return redirect(url_for('index'))

        filename = session_info['files'][company]
        filepath = session_dir / filename

        if not filepath.exists():
            flash('File not found.', 'error')
            return redirect(url_for('index'))

        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='text/markdown'
        )

    except Exception as e:
        flash(f'Download failed: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/download-all/<session_id>')
def download_all(session_id):
    """Download all prompts as a zip file."""
    try:
        session_dir = TEMP_DIR / session_id
        session_file = session_dir / 'session.json'

        if not session_file.exists():
            flash('Download session expired. Please generate prompts again.', 'error')
            return redirect(url_for('index'))

        session_info = json.loads(session_file.read_text(encoding='utf-8'))

        # Create zip file in memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for company, filename in session_info['files'].items():
                filepath = session_dir / filename
                if filepath.exists():
                    zip_file.write(filepath, filename)

        zip_buffer.seek(0)

        timestamp = session_info['timestamp'].replace(':', '').replace('-', '').replace(' ', '_')
        zip_filename = f"job_prompts_{timestamp}.zip"

        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )

    except Exception as e:
        flash(f'Download failed: {str(e)}', 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)